import openpyxl


# We are fetching the data from excel and storing it to the dict
class Testdataexcel:

    @staticmethod
    def data_excel(testcase):

        dict = {}  # Empty Dictonary

        workbook = openpyxl.load_workbook("C://Users//digvijayt//PycharmProjects//Py_Sel_Framework//TestData//TestData.xlsx")
        sheet = workbook.active

        for i in range(1, sheet.max_row + 1):  # +1 because it will return max-1
            if sheet.cell(row=i,
                          column=1).value == testcase:  # If you want to fetch only one row you can apply condition like this
                for j in range(1, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[dict]
